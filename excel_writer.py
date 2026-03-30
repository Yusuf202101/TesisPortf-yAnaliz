import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
DATA_FONT   = Font(name="Arial", size=9)
CENTER      = Alignment(horizontal="center", vertical="center")
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")


def build_excel(facility_data: dict, time_index: list) -> bytes:
    """
    facility_data = {
        tesis_name: {
            "kudüp": {ts: val, ...},
            "kgüp":  {ts: val, ...},
            "uevm":  {ts: val, ...},
        }
    }
    Dönüş: Excel dosyasının bytes içeriği (Streamlit download_button için)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Veriler"

    # Başlık — A1
    ws.column_dimensions["A"].width = 20
    c = ws["A1"]
    c.value = "Tarih-Saat"
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = CENTER; c.border = BORDER

    col = 2
    col_map = {}
    for tesis in facility_data:
        for vtype in ["kudüp", "kgüp", "uevm"]:
            ltr    = get_column_letter(col)
            header = f"{tesis} - {vtype.upper()}"
            cell   = ws[f"{ltr}1"]
            cell.value = header
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL
            cell.alignment = CENTER; cell.border = BORDER
            ws.column_dimensions[ltr].width = max(len(header) * 1.05, 14)
            col_map[(tesis, vtype)] = col
            col += 1

    # Veri satırları
    for row_idx, ts in enumerate(time_index, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        c = ws.cell(row=row_idx, column=1, value=ts)
        c.font = DATA_FONT; c.alignment = CENTER; c.border = BORDER
        if fill:
            c.fill = fill

        for (tesis, vtype), cidx in col_map.items():
            val  = facility_data[tesis][vtype].get(ts, "")
            cell = ws.cell(row=row_idx, column=cidx, value=val)
            cell.font = DATA_FONT; cell.alignment = CENTER; cell.border = BORDER
            if fill:
                cell.fill = fill

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
